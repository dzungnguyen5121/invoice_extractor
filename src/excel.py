import openpyxl
import os
import logging

def save_to_excel(data_list, excel_path='output/invoices.xlsx'):
    try:
        if os.path.exists(excel_path):
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(['STT', 'Số hóa đơn', 'Ngày tháng năm', 'Tên công ty', 'Mã số thuế', 'Tên hàng',
                       'Đơn vị', 'Số lượng', 'Đơn giá', 'Thành tiền', '% Thuế', 'Tiền thuế', 'Tổng tiền', 'Phân loại'])
        for data in data_list:
            row = [
                ws.max_row, data['invoice_number'], data['date'], data['company_name'], data['tax_code'],
                data['item_name'], data['unit'], data['quantity'], data['unit_price'], data['subtotal'],
                data['tax_rate'] * 100, data['tax_amount'], data['total'], data['category']
            ]
            ws.append(row)
        wb.save(excel_path)
        logging.info(f"Saved data to {excel_path}")
    except Exception as e:
        logging.error(f"Failed to save to Excel: {e}")
        raise 